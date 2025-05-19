import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

def analyze_and_flag_low_scores():
    # Set input file path
    input_file = "/Users/saimanpokhrel/Desktop/Data/Processed/final_summary_data.xlsx"
    
    # Get the directory from the input file path
    input_dir = os.path.dirname(input_file)
    
    # Construct output file path in the same directory
    output_file = os.path.join(input_dir, "final_data.xlsx")
    
    # Check if input file exists
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file '{input_file}' not found")
    
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active  # Get the first sheet

    # Identify column indices for key metrics
    header_row = 1
    columns = {}
    for cell in sheet[header_row]:
        if cell.value == "Class":
            columns["Class"] = cell.column - 1  # Convert to 0-based index
        elif cell.value == "Chapter":
            columns["Chapter"] = cell.column - 1
        elif cell.value == "LO_Mean":
            columns["LO_Mean"] = cell.column - 1
        elif cell.value == "Curiosity_Mean":
            columns["Curiosity_Mean"] = cell.column - 1
        elif cell.value == "Alignment_Mean":
            columns["Alignment_Mean"] = cell.column - 1

    # Verify all required columns exist
    required_columns = ["Class", "Chapter", "LO_Mean", "Curiosity_Mean", "Alignment_Mean"]
    for col in required_columns:
        if col not in columns:
            raise ValueError(f"Required column '{col}' not found in the sheet")

    # Find chapters with any metric below 4
    flagged_chapters = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        metrics = {
            "Class": row[columns["Class"]],
            "Chapter": row[columns["Chapter"]],
            "LO_Mean": row[columns["LO_Mean"]],
            "Curiosity_Mean": row[columns["Curiosity_Mean"]],
            "Alignment_Mean": row[columns["Alignment_Mean"]]
        }
        
        # Check if any metric is below 4 (only for numeric values)
        if any(isinstance(metrics[key], (int, float)) and metrics[key] < 4 
           for key in ["LO_Mean", "Curiosity_Mean", "Alignment_Mean"]):
            flagged_chapters.append(metrics)

    # Create summary sheet
    if "Flagged Chapters" in wb.sheetnames:
        del wb["Flagged Chapters"]
    summary_sheet = wb.create_sheet("Flagged Chapters")

    # Write headers with styling
    headers = ["Class", "Chapter", "LO_Mean", "Curiosity_Mean", "Alignment_Mean", "Issues"]
    summary_sheet.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Write flagged chapters
    red_font = Font(color="FF0000")
    for chapter in flagged_chapters:
        issues = []
        for metric in ["LO_Mean", "Curiosity_Mean", "Alignment_Mean"]:
            if isinstance(chapter[metric], (int, float)) and chapter[metric] < 4:
                issues.append(metric)
        
        row = [
            chapter["Class"],
            chapter["Chapter"],
            chapter["LO_Mean"],
            chapter["Curiosity_Mean"],
            chapter["Alignment_Mean"],
            ", ".join(issues)
        ]
        summary_sheet.append(row)

        # Highlight problematic cells
        for col_idx, metric in enumerate(["LO_Mean", "Curiosity_Mean", "Alignment_Mean"], start=3):
            if metric in issues:
                summary_sheet.cell(row=summary_sheet.max_row, column=col_idx).font = red_font

    # Adjust column widths
    for col in summary_sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        summary_sheet.column_dimensions[col_letter].width = max_length + 2

    # Save the new file in the same directory as input
    wb.save(output_file)
    print(f"Analysis complete. Output saved to '{output_file}'")
    print(f"Found {len(flagged_chapters)} chapters needing improvement")

if __name__ == "__main__":
    analyze_and_flag_low_scores()
